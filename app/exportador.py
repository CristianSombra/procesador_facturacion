import pandas as pd
from tkinter import filedialog
from reportlab.lib.pagesizes import landscape, A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from database import get_connection


COLUMNAS_EXPORTACION = {
    "tipo_factura": "Tipo de comprobante",
    "numero_factura": "Número",
    "fecha_facturacion": "Fecha de facturación",
    "fecha_vencimiento": "Fecha de vencimiento",
    "razon_social": "Razón social",
    "cuit": "CUIT",
    "condicion_iva": "Condición IVA",
    "domicilio": "Domicilio",
    "condicion_venta": "Condición de venta",
    "periodo": "Período facturado",
    "importe_total": "Importe total",
    "cae": "CAE",
}

def formatear_importe(valor):
    try:
        valor = float(valor)
        return f"$ {valor:,.0f}".replace(",", ".")
    except:
        return valor


def obtener_datos():
    conn = get_connection()

    query = """
    SELECT
        tipo_factura,
        numero_factura,
        fecha_facturacion,
        fecha_vencimiento,
        razon_social,
        cuit,
        condicion_iva,
        domicilio,
        condicion_venta,
        periodo,
        importe_total,
        cae
    FROM facturas
    ORDER BY fecha_facturacion, numero_factura;
    """

    df = pd.read_sql(query, conn)
    conn.close()

    df = df.rename(columns=COLUMNAS_EXPORTACION)
    return df


def exportar_excel():
    df = obtener_datos()

    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivo Excel", "*.xlsx")],
        title="Guardar Excel"
    )

    if not ruta:
        return

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Facturación")

        ws = writer.sheets["Facturación"]

        for celda in ws[1]:
            celda.font = Font(bold=True)

        col_importe = None

        for idx, celda in enumerate(ws[1], start=1):
            if celda.value == "Importe total":
                col_importe = idx
                break

        if col_importe:
            for fila in range(2, ws.max_row + 1):
                celda = ws.cell(row=fila, column=col_importe)
                celda.number_format = '$ #,##0'

        for col in ws.columns:
            letra = get_column_letter(col[0].column)
            max_largo = 0

            for celda in col:
                if celda.value is not None:
                    max_largo = max(max_largo, len(str(celda.value)))

            ws.column_dimensions[letra].width = min(max_largo + 3, 55)


def exportar_pdf():
    df = obtener_datos()

    df_pdf = df.copy()
    df_pdf["Importe total"] = df_pdf["Importe total"].apply(formatear_importe)

    ruta = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("Archivo PDF", "*.pdf")],
        title="Guardar PDF"
    )

    if not ruta:
        return

    documento = SimpleDocTemplate(
        ruta,
        pagesize=landscape(A3),
        rightMargin=8,
        leftMargin=8,
        topMargin=15,
        bottomMargin=15
    )

    estilos = getSampleStyleSheet()
    estilo_celda = ParagraphStyle(
        "celda",
        fontName="Helvetica",
        fontSize=5,
        leading=6
    )

    estilo_header = ParagraphStyle(
        "header",
        fontName="Helvetica-Bold",
        fontSize=5,
        leading=6
    )

    elementos = [
        Paragraph("Base de Datos de Facturación", estilos["Title"]),
        Spacer(1, 10)
    ]

    encabezados = [Paragraph(str(col), estilo_header) for col in df_pdf.columns]

    filas = []
    for _, fila in df_pdf.iterrows():
        filas.append([
            Paragraph(str(valor), estilo_celda)
            for valor in fila.values
        ])

    datos_tabla = [encabezados] + filas

    anchos = [
        75,   # Tipo comprobante
        55,   # Número
        60,   # Fecha facturación
        60,   # Fecha vencimiento
        175,  # Razón social
        70,   # CUIT
        80,   # Condición IVA
        170,  # Domicilio
        115,  # Condición venta
        95,   # Período
        65,   # Importe
        80,   # CAE
    ]

    tabla = Table(
        datos_tabla,
        colWidths=anchos,
        repeatRows=1
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    elementos.append(tabla)
    documento.build(elementos)